import pandas as pd
from datetime import datetime, date
import io
from typing import Optional
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils.dataframe import dataframe_to_rows

class ExcelExporter:
    """Classe pour exporter les données financières vers Excel"""
    
    def __init__(self, data_manager):
        self.data_manager = data_manager
    
    def generate_monthly_report(self, year: int, month: int) -> bytes:
        """Génère un rapport mensuel complet en Excel"""
        
        # Créer un nouveau classeur
        wb = openpyxl.Workbook()
        
        # Supprimer la feuille par défaut
        wb.remove(wb.active)
        
        # Charger les données
        transactions_df = self.data_manager.load_transactions()
        employees_df = self.data_manager.load_employees()
        shareholders_df = self.data_manager.load_shareholders()
        
        # Filtrer les transactions du mois
        if not transactions_df.empty:
            transactions_df['date'] = pd.to_datetime(transactions_df['date'])
            monthly_transactions = transactions_df[
                (transactions_df['date'].dt.year == year) & 
                (transactions_df['date'].dt.month == month)
            ].copy()
        else:
            monthly_transactions = pd.DataFrame()
        
        # Créer les feuilles
        self._create_summary_sheet(wb, monthly_transactions, employees_df, shareholders_df, year, month)
        self._create_transactions_sheet(wb, monthly_transactions)
        self._create_salaries_sheet(wb, employees_df, monthly_transactions, year, month)
        self._create_profit_sharing_sheet(wb, shareholders_df, monthly_transactions)
        
        # Sauvegarder dans un buffer
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
    
    def _create_summary_sheet(self, wb, transactions_df, employees_df, shareholders_df, year, month):
        """Crée la feuille de résumé"""
        ws = wb.create_sheet("Résumé")
        
        # Titre
        ws['A1'] = f"Rapport Financier - {month:02d}/{year}"
        ws['A1'].font = Font(size=16, bold=True)
        ws.merge_cells('A1:D1')
        
        # Période
        ws['A3'] = "Période:"
        ws['B3'] = f"{month:02d}/{year}"
        ws['A3'].font = Font(bold=True)
        
        # Date de génération
        ws['A4'] = "Généré le:"
        ws['B4'] = datetime.now().strftime("%d/%m/%Y à %H:%M")
        ws['A4'].font = Font(bold=True)
        
        # Résumé financier
        ws['A6'] = "RÉSUMÉ FINANCIER"
        ws['A6'].font = Font(size=14, bold=True)
        
        if not transactions_df.empty:
            total_income = transactions_df[transactions_df['type'] == 'Entrée']['montant'].sum()
            total_expenses = transactions_df[transactions_df['type'] == 'Sortie']['montant'].sum()
        else:
            total_income = 0
            total_expenses = 0
        
        net_result = total_income - total_expenses
        
        ws['A8'] = "Total Entrées:"
        ws['B8'] = total_income
        ws['A9'] = "Total Sorties:"
        ws['B9'] = total_expenses
        ws['A10'] = "Résultat Net:"
        ws['B10'] = net_result
        
        # Formatage des cellules financières
        for row in range(8, 11):
            ws[f'B{row}'].number_format = '#,##0.00 €'
            if row == 10:  # Résultat net
                ws[f'A{row}'].font = Font(bold=True)
                ws[f'B{row}'].font = Font(bold=True)
        
        # Informations sur les employés
        ws['A12'] = "INFORMATIONS EMPLOYÉS"
        ws['A12'].font = Font(size=14, bold=True)
        
        if not employees_df.empty:
            active_employees = employees_df[employees_df['actif'] == True]
            total_monthly_salaries = active_employees['salaire_mensuel'].sum()
            
            ws['A14'] = "Nombre d'employés actifs:"
            ws['B14'] = len(active_employees)
            ws['A15'] = "Total salaires mensuels:"
            ws['B15'] = total_monthly_salaries
            ws['B15'].number_format = '#,##0.00 €'
        
        # Informations sur les actionnaires
        ws['A17'] = "INFORMATIONS ACTIONNAIRES"
        ws['A17'].font = Font(size=14, bold=True)
        
        if not shareholders_df.empty:
            active_shareholders = shareholders_df[shareholders_df['actif'] == True]
            
            ws['A19'] = "Nombre d'actionnaires actifs:"
            ws['B19'] = len(active_shareholders)
            
            if net_result > 0:
                ws['A20'] = "Bénéfices disponibles:"
                ws['B20'] = max(0, net_result)
                ws['B20'].number_format = '#,##0.00 €'
        
        # Ajuster la largeur des colonnes
        ws.column_dimensions['A'].width = 25
        ws.column_dimensions['B'].width = 15
    
    def _create_transactions_sheet(self, wb, transactions_df):
        """Crée la feuille des transactions"""
        ws = wb.create_sheet("Transactions")
        
        ws['A1'] = "LISTE DES TRANSACTIONS"
        ws['A1'].font = Font(size=14, bold=True)
        
        if not transactions_df.empty:
            # Préparer les données
            display_df = transactions_df.copy()
            display_df['date'] = display_df['date'].dt.strftime('%d/%m/%Y')
            display_df = display_df[['date', 'type', 'description', 'montant', 'categorie']]
            display_df.columns = ['Date', 'Type', 'Description', 'Montant', 'Catégorie']
            
            # Ajouter les en-têtes
            for col_num, column_title in enumerate(display_df.columns, 1):
                cell = ws.cell(row=3, column=col_num)
                cell.value = column_title
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            # Ajouter les données
            for row_num, (_, row_data) in enumerate(display_df.iterrows(), 4):
                for col_num, value in enumerate(row_data, 1):
                    cell = ws.cell(row=row_num, column=col_num)
                    cell.value = value
                    if col_num == 4:  # Colonne montant
                        cell.number_format = '#,##0.00 €'
            
            # Ajuster les largeurs
            ws.column_dimensions['A'].width = 12
            ws.column_dimensions['B'].width = 10
            ws.column_dimensions['C'].width = 30
            ws.column_dimensions['D'].width = 15
            ws.column_dimensions['E'].width = 15
        else:
            ws['A3'] = "Aucune transaction pour cette période"
    
    def _create_salaries_sheet(self, wb, employees_df, transactions_df, year, month):
        """Crée la feuille des salaires"""
        ws = wb.create_sheet("Salaires")
        
        ws['A1'] = f"GESTION DES SALAIRES - {month:02d}/{year}"
        ws['A1'].font = Font(size=14, bold=True)
        
        if not employees_df.empty:
            # En-têtes
            headers = ['Employé', 'Poste', 'Salaire Mensuel', 'Payé ce Mois', 'Reste à Payer']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=3, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            # Données des employés
            for row_num, (_, employee) in enumerate(employees_df.iterrows(), 4):
                if employee['actif']:
                    # Calculer les paiements effectués ce mois
                    if not transactions_df.empty:
                        payments = transactions_df[
                            transactions_df['description'].str.contains(f"Salaire - {employee['nom']}", na=False)
                        ]['montant'].sum()
                    else:
                        payments = 0
                    
                    remaining = employee['salaire_mensuel'] - payments
                    
                    ws.cell(row=row_num, column=1, value=f"{employee['prenom']} {employee['nom']}")
                    ws.cell(row=row_num, column=2, value=employee['poste'])
                    ws.cell(row=row_num, column=3, value=employee['salaire_mensuel'])
                    ws.cell(row=row_num, column=4, value=payments)
                    ws.cell(row=row_num, column=5, value=max(0, remaining))
                    
                    # Formatage des montants
                    for col in [3, 4, 5]:
                        ws.cell(row=row_num, column=col).number_format = '#,##0.00 €'
            
            # Ajuster les largeurs
            for col, width in enumerate([20, 15, 15, 15, 15], 1):
                ws.column_dimensions[chr(64 + col)].width = width
        else:
            ws['A3'] = "Aucun employé enregistré"
    
    def _create_profit_sharing_sheet(self, wb, shareholders_df, transactions_df):
        """Crée la feuille de partage des bénéfices"""
        ws = wb.create_sheet("Partage Bénéfices")
        
        ws['A1'] = "PARTAGE DES BÉNÉFICES"
        ws['A1'].font = Font(size=14, bold=True)
        
        if not shareholders_df.empty:
            # Calculer le bénéfice net
            if not transactions_df.empty:
                total_income = transactions_df[transactions_df['type'] == 'Entrée']['montant'].sum()
                total_expenses = transactions_df[transactions_df['type'] == 'Sortie']['montant'].sum()
                net_profit = total_income - total_expenses
            else:
                net_profit = 0
            
            ws['A3'] = "Bénéfice Net à Distribuer:"
            ws['B3'] = max(0, net_profit)
            ws['B3'].number_format = '#,##0.00 €'
            ws['A3'].font = Font(bold=True)
            
            # En-têtes
            headers = ['Actionnaire', '% Actions', 'Part du Bénéfice']
            for col_num, header in enumerate(headers, 1):
                cell = ws.cell(row=5, column=col_num)
                cell.value = header
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
            
            # Données des actionnaires
            total_percentage = 0
            for row_num, (_, shareholder) in enumerate(shareholders_df.iterrows(), 6):
                if shareholder['actif']:
                    share_amount = (shareholder['pourcentage_actions'] / 100) * max(0, net_profit)
                    total_percentage += shareholder['pourcentage_actions']
                    
                    ws.cell(row=row_num, column=1, value=f"{shareholder['prenom']} {shareholder['nom']}")
                    ws.cell(row=row_num, column=2, value=f"{shareholder['pourcentage_actions']}%")
                    ws.cell(row=row_num, column=3, value=share_amount)
                    
                    ws.cell(row=row_num, column=3).number_format = '#,##0.00 €'
            
            # Total des pourcentages
            last_row = len(shareholders_df[shareholders_df['actif'] == True]) + 6
            ws.cell(row=last_row + 1, column=1, value="TOTAL").font = Font(bold=True)
            ws.cell(row=last_row + 1, column=2, value=f"{total_percentage}%").font = Font(bold=True)
            
            # Ajuster les largeurs
            ws.column_dimensions['A'].width = 20
            ws.column_dimensions['B'].width = 12
            ws.column_dimensions['C'].width = 18
        else:
            ws['A3'] = "Aucun actionnaire enregistré"
    
    def export_all_data(self) -> bytes:
        """Exporte toutes les données vers Excel"""
        wb = openpyxl.Workbook()
        wb.remove(wb.active)
        
        # Charger toutes les données
        transactions_df = self.data_manager.load_transactions()
        employees_df = self.data_manager.load_employees()
        shareholders_df = self.data_manager.load_shareholders()
        
        # Créer les feuilles pour toutes les données
        self._export_all_transactions(wb, transactions_df)
        self._export_all_employees(wb, employees_df)
        self._export_all_shareholders(wb, shareholders_df)
        
        # Sauvegarder
        buffer = io.BytesIO()
        wb.save(buffer)
        buffer.seek(0)
        
        return buffer.getvalue()
    
    def _export_all_transactions(self, wb, transactions_df):
        """Exporte toutes les transactions"""
        ws = wb.create_sheet("Toutes les Transactions")
        
        if not transactions_df.empty:
            display_df = transactions_df.copy()
            if 'date' in display_df.columns:
                display_df['date'] = pd.to_datetime(display_df['date']).dt.strftime('%d/%m/%Y')
            
            # Ajouter les données avec en-têtes
            for r in dataframe_to_rows(display_df, index=False, header=True):
                ws.append(r)
            
            # Formatage des en-têtes
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
        else:
            ws['A1'] = "Aucune transaction"
    
    def _export_all_employees(self, wb, employees_df):
        """Exporte tous les employés"""
        ws = wb.create_sheet("Tous les Employés")
        
        if not employees_df.empty:
            display_df = employees_df.copy()
            if 'date_embauche' in display_df.columns:
                display_df['date_embauche'] = pd.to_datetime(display_df['date_embauche']).dt.strftime('%d/%m/%Y')
            
            for r in dataframe_to_rows(display_df, index=False, header=True):
                ws.append(r)
            
            # Formatage des en-têtes
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
        else:
            ws['A1'] = "Aucun employé"
    
    def _export_all_shareholders(self, wb, shareholders_df):
        """Exporte tous les actionnaires"""
        ws = wb.create_sheet("Tous les Actionnaires")
        
        if not shareholders_df.empty:
            for r in dataframe_to_rows(shareholders_df, index=False, header=True):
                ws.append(r)
            
            # Formatage des en-têtes
            for cell in ws[1]:
                cell.font = Font(bold=True)
                cell.fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
                cell.font = Font(color="FFFFFF", bold=True)
        else:
            ws['A1'] = "Aucun actionnaire"
